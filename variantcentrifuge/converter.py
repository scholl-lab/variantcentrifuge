# File: variantcentrifuge/converter.py
# Location: variantcentrifuge/variantcentrifuge/converter.py

"""
File conversion module.

This module provides functionality to convert TSV files to XLSX format
and append additional sheets. It now also supports producing JSON files
for the HTML report.
"""

import hashlib
import json
import logging
import os
import re
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger("variantcentrifuge")


def convert_to_excel(tsv_file: str, cfg: Dict[str, Any]) -> str:
    """
    Convert a TSV file to XLSX format with a single "Results" sheet.

    Parameters
    ----------
    tsv_file : str
        Path to the TSV file.
    cfg : dict
        Configuration dictionary containing various settings.

    Returns
    -------
    str
        The path to the generated XLSX file.
    """
    # MODIFIED: Start of empty report generation
    # Read the TSV, handling the case where it might only have a header
    # Define appropriate dtypes for genomic data to avoid mixed type warnings
    genomic_dtypes = {
        "CHROM": "str",
        "POS": "str",  # Use string to handle both numeric positions and special values like "."
        "REF": "str",
        "ALT": "str",
        "QUAL": "str",  # Quality scores can be "." or numeric
        "FILTER": "str",
    }

    # Read with low_memory=False to handle mixed types gracefully and suppress warnings
    df = pd.read_csv(tsv_file, sep="\t", na_values="NA", dtype=genomic_dtypes, low_memory=False)

    # Log whether we have data or just headers
    if len(df) == 0:
        logger.warning("TSV file contains only headers or is empty. Creating header-only Excel.")

    # Remove raw igv_links column if it exists to avoid duplication with
    # the formatted hyperlinked IGV Report Links column added in finalize_excel_file
    if "igv_links" in df.columns:
        logger.debug("Removing raw igv_links column from Excel conversion")
        df = df.drop(columns=["igv_links"])

    xlsx_file = os.path.splitext(tsv_file)[0] + ".xlsx"
    df.to_excel(xlsx_file, index=False, sheet_name="Results")
    # MODIFIED: End of empty report generation
    return xlsx_file


def append_tsv_as_sheet(xlsx_file: str, tsv_file: str, sheet_name: str = "Metadata") -> None:
    """
    Append a TSV file as a new sheet to an existing XLSX file.

    This function reads the TSV file and appends it as a new sheet in the existing
    XLSX file. Assumes the TSV file has a header row.

    Parameters
    ----------
    xlsx_file : str
        Path to the existing XLSX file.
    tsv_file : str
        Path to the TSV file to append as a sheet.
    sheet_name : str, optional
        Name of the new sheet to be appended. Defaults to "Metadata".

    Returns
    -------
    None
    """
    # Read with low_memory=False to suppress dtype warnings for metadata files
    df = pd.read_csv(tsv_file, sep="\t", header=0, low_memory=False)
    with pd.ExcelWriter(xlsx_file, engine="openpyxl", mode="a", if_sheet_exists="new") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)


def finalize_excel_file(xlsx_file: str, cfg: Dict[str, Any]) -> None:
    """
    Apply final formatting to all sheets in xlsx_file.

    - Freeze the top row
    - Enable auto-filter
    - Only for the 'Results' sheet, generate hyperlinks from CHROM, POS, REF, ALT
      using cfg["links"] (the link template dictionary).
    - Add "IGV Report Links" column with hyperlinks to IGV reports if available

    Parameters
    ----------
    xlsx_file : str
        Path to the Excel file to finalize.
    cfg : dict
        Configuration dictionary containing links and IGV settings.

    Returns
    -------
    None
    """
    # Using load_workbook and get_column_letter already imported at module level
    wb = load_workbook(xlsx_file)
    link_templates = cfg.get("links", {})

    # Check for IGV reports map
    igv_enabled = cfg.get("igv_enabled", False)
    igv_map = None
    igv_lookup = {}

    if igv_enabled:
        # Log that we're looking for the IGV mapping file
        logger.info("Looking for IGV reports map for Excel integration...")
        output_dir = os.path.dirname(xlsx_file)
        igv_map_path = os.path.join(output_dir, "report", "igv", "igv_reports_map.json")
        if not os.path.exists(igv_map_path):
            # Try the parent directory as an alternative location
            parent_dir = os.path.dirname(os.path.dirname(output_dir))
            alt_path = os.path.join(parent_dir, "report", "igv", "igv_reports_map.json")
            if os.path.exists(alt_path):
                igv_map_path = alt_path

        if os.path.exists(igv_map_path):
            logger.info(f"Found IGV reports mapping file: {igv_map_path}")
            try:
                # Load IGV reports mapping
                with open(igv_map_path, "r", encoding="utf-8") as f:
                    igv_map_data = json.load(f)

                # Handle both old and new IGV map format
                if "variants" in igv_map_data:
                    # New format - variant-centric with nested sample reports
                    igv_map = igv_map_data["variants"]
                    # Count total entries (number of sample-variant combinations)
                    total_entries = sum(len(entry.get("sample_reports", {})) for entry in igv_map)
                    logger.info(
                        f"Loaded IGV reports map with {len(igv_map)} variants and "
                        f"{total_entries} total reports"
                    )
                else:
                    # Old format - flat list of entries
                    igv_map = igv_map_data
                    logger.info(f"Loaded IGV reports map with {len(igv_map)} entries")

                # MODIFIED: Start of IGV Excel link fix
                # Populate the igv_lookup dictionary for quick access when adding links to Excel
                igv_lookup = {}
                for entry in igv_map:
                    # Extract variant identifiers
                    chrom = entry.get("chrom", "")
                    pos = str(entry.get("pos", ""))
                    ref = entry.get("ref", "")
                    alt = entry.get("alt", "")

                    # Handle new format with sample_reports nested dictionary
                    if "sample_reports" in entry:
                        for sample_id, report_path in entry["sample_reports"].items():
                            key = (chrom, pos, ref, alt, sample_id)
                            igv_lookup[key] = report_path
                    # Handle old format with flat sample_id and report_path
                    elif "sample_id" in entry and "report_path" in entry:
                        key = (chrom, pos, ref, alt, entry["sample_id"])
                        igv_lookup[key] = entry["report_path"]

                logger.info(f"Created lookup for {len(igv_lookup)} IGV report entries")
                # MODIFIED: End of IGV Excel link fix
            except Exception as e:
                logger.error(f"Error loading IGV reports map: {e}")
                igv_map = None
        else:
            logger.warning(f"IGV reports map file not found at: {igv_map_path}")
            logger.warning("Excel file will not include IGV report links")

    for ws in wb.worksheets:
        # 1. Freeze top row
        ws.freeze_panes = "A2"

        # 2. Enable auto-filter
        max_col_letter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{max_col_letter}1"

        # Only generate links in the 'Results' sheet
        if ws.title != "Results":
            continue

        # 3. Generate links in the Results sheet
        # First find required column indices
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header_indices = {}
        for idx, col_name in enumerate(header_row, 1):  # 1-indexed for openpyxl
            if col_name in ["CHROM", "POS", "REF", "ALT", "GT"]:
                header_indices[col_name] = idx

        # Add IGV Report Links column if IGV enabled and map is available
        igv_col = None
        if igv_enabled and igv_map:
            # Add the IGV Report Links column header
            igv_col = ws.max_column + 1
            igv_col_letter = get_column_letter(igv_col)
            igv_header_cell = ws[f"{igv_col_letter}1"]
            igv_header_cell.value = "IGV Report Links"

        # Process data rows for both regular links and IGV links
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):  # Skip header
            # Extract variant identifiers from the row
            if all(col in header_indices for col in ["CHROM", "POS", "REF", "ALT"]):
                chrom = str(ws.cell(row=row_idx, column=header_indices["CHROM"]).value or "")
                pos = str(ws.cell(row=row_idx, column=header_indices["POS"]).value or "")
                ref = str(ws.cell(row=row_idx, column=header_indices["REF"]).value or "")
                alt = str(ws.cell(row=row_idx, column=header_indices["ALT"]).value or "")

                # Add IGV Report Links if enabled
                if igv_enabled and igv_map and igv_col and "GT" in header_indices:
                    # Get the GT value for this row
                    gt_cell = ws.cell(row=row_idx, column=header_indices["GT"])
                    gt_value = gt_cell.value or ""
                    gt_value = str(gt_value) if gt_value is not None else ""

                    if not gt_value:
                        # No GT value, set N/A
                        igv_cell = ws.cell(row=row_idx, column=igv_col)
                        igv_cell.value = "N/A"
                        continue

                    # Create a regular expression pattern for GT parsing
                    pattern = re.compile(r"([^()]+)\(([^)]+)\)")

                    # Find all related IGV reports for this variant
                    igv_reports = []
                    sample_entries = gt_value.split(";")
                    for entry in sample_entries:
                        entry = entry.strip()
                        if not entry:
                            continue

                        m = pattern.match(entry)
                        if m:
                            sample_id = m.group(1).strip()
                            genotype = m.group(2).strip()

                            # Skip reference genotypes
                            if genotype in ["0/0", "./."]:
                                continue

                            # Look up the IGV report path
                            lookup_key = (chrom, pos, ref, alt, sample_id)
                            if lookup_key in igv_lookup:
                                # Get path from lookup (e.g., "igv/S1_var.html")
                                original_path_from_map = igv_lookup[lookup_key]
                                igv_reports.append((sample_id, original_path_from_map))

                    # Update the IGV links cell
                    igv_cell = ws.cell(row=row_idx, column=igv_col)
                    if len(igv_reports) == 1:
                        # Single report - make a hyperlink
                        sample_id, original_path_from_map = igv_reports[0]
                        # Make sure the cell value is sortable
                        igv_cell.value = sample_id  # This is what will be used for sorting
                        excel_relative_hyperlink_path = os.path.join(
                            "report", original_path_from_map
                        )
                        igv_cell.hyperlink = excel_relative_hyperlink_path
                        igv_cell.style = "Hyperlink"
                    elif len(igv_reports) > 1:
                        # Multiple reports - for sortability use first sample ID as value
                        first_sample = igv_reports[0][0]  # First sample ID
                        link_texts = []
                        for sid, rpath in igv_reports:
                            link_texts.append(f"{sid} ({os.path.join('report', rpath)})")
                        # Use first sample ID as sortable value, but show all in the tooltip/display
                        # MODIFIED: Make display format more informative
                        num_others = len(igv_reports) - 1
                        igv_cell.value = f"{first_sample} (+{num_others} others)"
                        # Add comment with full details for all reports
                        from openpyxl.comments import Comment

                        igv_cell.comment = Comment("; ".join(link_texts), "IGV Reports")
                    else:
                        igv_cell.value = "N/A"

                # We're not adding the regular external links from cfg["links"] anymore
                # because they're redundant with the IGV Report Links column
                # and causing confusion in the Excel report
                # The HTML report still shows all links as configured

    wb.save(xlsx_file)


def produce_report_json(variant_tsv: str, output_dir: str) -> None:
    """
    Produce JSON files (variants.json and summary.json) from a TSV of variants.

    The TSV file is expected to have columns including GENE, CHROM, POS, REF, ALT, IMPACT.
    Two JSON files are produced:
    
    - variants.json: A list of all variants with their annotations.
    - summary.json: Aggregated counts (num_variants, num_genes, impact_distribution).
    
    These files are used by the HTML report generator.

    Parameters
    ----------
    variant_tsv : str
        Path to the TSV file containing annotated variants.
    output_dir : str
        Directory to write the JSON outputs (variants.json, summary.json).

    Returns
    -------
    None
    """
    logger.info(f"Producing JSON files for HTML report from {variant_tsv}")

    # MODIFIED: Start of empty report generation
    # Read the variants TSV, handling the case where it might only have a header
    try:
        # Use low_memory=False to suppress dtype warnings for variant files
        df = pd.read_csv(variant_tsv, sep="\t", low_memory=False)
    except pd.errors.EmptyDataError:
        logger.warning("Empty TSV file provided. Generating empty JSON files.")
        df = pd.DataFrame()
    # MODIFIED: End of empty report generation

    # Convert to list of dictionaries for JSON
    variants = df.to_dict(orient="records") if not df.empty else []

    # Sanitize the variants list to ensure JSON compliance
    for variant_dict in variants:
        for key, val in variant_dict.items():
            if pd.isna(val):
                variant_dict[key] = None

    # Check for IGV reports mapping file
    igv_map_path = os.path.join(output_dir, "report", "igv", "igv_reports_map.json")
    if os.path.exists(igv_map_path):
        logger.info(f"Found IGV reports mapping file: {igv_map_path}")
        try:
            # Load IGV reports mapping
            with open(igv_map_path, "r", encoding="utf-8") as f:
                igv_map_data = json.load(f)

            # Handle both old and new IGV map format
            if "variants" in igv_map_data:
                # New format - variant-centric with nested sample reports
                igv_map = igv_map_data["variants"]
            else:
                # Old format - flat list of entries
                igv_map = igv_map_data

            # Create an efficient lookup dictionary
            # Key: (chrom, pos, ref, alt, sample_id), Value: report_path
            igv_lookup = {}
            for entry in igv_map:
                chrom = entry["chrom"]
                pos = str(entry["pos"])
                ref = entry["ref"]
                alt = entry["alt"]

                # Handle new format with sample_reports nested dictionary
                if "sample_reports" in entry:
                    for sample_id, report_path in entry["sample_reports"].items():
                        key = (chrom, pos, ref, alt, sample_id)
                        igv_lookup[key] = report_path
                # Handle old format with flat sample_id and report_path
                elif "sample_id" in entry and "report_path" in entry:
                    key = (chrom, pos, ref, alt, entry["sample_id"])
                    igv_lookup[key] = entry["report_path"]

            # Pattern to extract (SampleID, GenotypeString) tuples from GT column
            pattern = re.compile(r"([^()]+)\(([^)]+)\)")

            # Augment variants with IGV links
            for variant_dict in variants:
                # Initialize igv_links list
                variant_dict["igv_links"] = []

                # Extract variant identifiers
                chrom = str(variant_dict.get("CHROM", ""))
                pos = str(variant_dict.get("POS", ""))
                ref = str(variant_dict.get("REF", ""))
                alt = str(variant_dict.get("ALT", ""))

                # Extract sample IDs with non-reference genotypes
                gt_value = variant_dict.get("GT", "")
                if gt_value:
                    sample_entries = gt_value.split(";")
                    for entry in sample_entries:
                        entry = entry.strip()
                        if not entry:
                            continue

                        m = pattern.match(entry)
                        if m:
                            sample_id = m.group(1).strip()
                            genotype = m.group(2).strip()

                            # Skip reference genotypes
                            if genotype in ["0/0", "./."]:
                                continue

                            # Look up the IGV report path
                            lookup_key = (chrom, pos, ref, alt, sample_id)
                            if lookup_key in igv_lookup:
                                variant_dict["igv_links"].append(
                                    {"sample_id": sample_id, "report_path": igv_lookup[lookup_key]}
                                )

            logger.info("Enriched variants with IGV report links")
        except Exception as e:
            logger.error(f"Failed to process IGV reports map: {str(e)}")
            # Continue with regular JSON generation without IGV links

    # MODIFIED: Start of empty report generation
    # Create summary information, handling the empty case properly
    num_variants = len(df) if not df.empty else 0
    unique_genes = df["GENE"].unique() if not df.empty and "GENE" in df.columns else []
    num_genes = len(unique_genes)

    impact_counts = {}
    if not df.empty and "IMPACT" in df.columns:
        impact_counts = df["IMPACT"].value_counts().to_dict()
    # MODIFIED: End of empty report generation

    summary_data = {
        "num_variants": num_variants,
        "num_genes": num_genes,
        "impact_distribution": impact_counts,
    }

    # Ensure the report directory exists
    variants_json_path = os.path.join(output_dir, "report", "variants.json")
    os.makedirs(os.path.dirname(variants_json_path), exist_ok=True)

    # Write the enriched variants JSON
    with open(variants_json_path, "w", encoding="utf-8") as vjf:
        json.dump(variants, vjf, indent=2)

    summary_json_path = os.path.join(output_dir, "report", "summary.json")
    with open(summary_json_path, "w", encoding="utf-8") as sjf:
        json.dump(summary_data, sjf, indent=2)

    logger.info(f"JSON files produced: {variants_json_path}, {summary_json_path}")
