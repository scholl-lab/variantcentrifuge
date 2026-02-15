"""
Full Excel output fidelity tests.

Comprehensive end-to-end tests verifying that optimized Excel generation
(xlsxwriter + openpyxl two-pass) produces functionally equivalent output
with all features:
- Multiple sheets (Results, Metadata, Statistics, Gene Burden)
- Freeze panes on all sheets
- Auto-filters on all sheets
- URL hyperlinks with proper styling
- GT cache columns excluded from output
"""

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from variantcentrifuge.converter import (
    append_tsv_as_sheet,
    convert_to_excel,
    finalize_excel_file,
)


def _create_test_variants_df(n_variants: int = 20, n_samples: int = 3, seed: int = 42):
    """Generate synthetic variant DataFrame for testing."""
    rng = np.random.default_rng(seed)

    # Generate genomic data
    chroms = rng.choice(["chr1", "chr2", "chr3"], size=n_variants)
    positions = rng.integers(1000, 100000, size=n_variants)
    refs = rng.choice(["A", "C", "G", "T"], size=n_variants)
    alts = rng.choice(["A", "C", "G", "T"], size=n_variants)

    # Generate GT column
    gts = []
    for _ in range(n_variants):
        sample_gts = [
            f"Sample{i+1}({rng.choice(['0/0', '0/1', '1/1', './.'])})" for i in range(n_samples)
        ]
        gts.append(";".join(sample_gts))

    # Generate annotations and URL columns
    genes = rng.choice(["BRCA1", "TP53", "EGFR"], size=n_variants)
    impacts = rng.choice(["HIGH", "MODERATE", "LOW"], size=n_variants)

    # Add URL columns for hyperlink testing
    splicai_urls = [
        f"https://spliceailookup.broadinstitute.org/#variant={chrom}-{pos}-{ref}-{alt}"
        for chrom, pos, ref, alt in zip(chroms, positions, refs, alts, strict=True)
    ]
    gnomad_urls = [
        f"https://gnomad.broadinstitute.org/variant/{chrom}-{pos}-{ref}-{alt}"
        for chrom, pos, ref, alt in zip(chroms, positions, refs, alts, strict=True)
    ]

    return pd.DataFrame(
        {
            "CHROM": chroms,
            "POS": positions,
            "REF": refs,
            "ALT": alts,
            "GT": gts,
            "GENE": genes,
            "IMPACT": impacts,
            "SpliceAI_URL": splicai_urls,
            "gnomAD_URL": gnomad_urls,
        }
    )


@pytest.mark.unit
def test_full_excel_with_all_sheets(tmp_path):
    """
    Verify Excel generation with all 4 sheets: Results, Metadata, Statistics, Gene Burden.

    Tests end-to-end flow:
    1. Create Results sheet via convert_to_excel
    2. Append Metadata, Statistics, Gene Burden sheets
    3. Finalize with formatting
    4. Verify all sheets exist with correct data
    """
    # Create test data
    variants_df = _create_test_variants_df(n_variants=20)
    metadata_df = pd.DataFrame(
        {"Key": ["sample_count", "gene_count"], "Value": ["3", "3"]}
    )
    stats_df = pd.DataFrame(
        {"Category": ["HIGH", "MODERATE", "LOW"], "Count": [5, 10, 5]}
    )
    gene_burden_df = pd.DataFrame(
        {"Gene": ["BRCA1", "TP53", "EGFR"], "Variant_Count": [7, 8, 5]}
    )

    # Write TSV files
    variants_tsv = tmp_path / "variants.tsv"
    metadata_tsv = tmp_path / "metadata.tsv"
    stats_tsv = tmp_path / "statistics.tsv"
    gene_burden_tsv = tmp_path / "gene_burden.tsv"

    variants_df.to_csv(variants_tsv, sep="\t", index=False)
    metadata_df.to_csv(metadata_tsv, sep="\t", index=False)
    stats_df.to_csv(stats_tsv, sep="\t", index=False)
    gene_burden_df.to_csv(gene_burden_tsv, sep="\t", index=False)

    # Generate Excel with all sheets
    cfg = {"links": {}, "igv_enabled": False}
    xlsx_path = convert_to_excel(str(variants_tsv), cfg, df=variants_df)
    append_tsv_as_sheet(xlsx_path, str(metadata_tsv), sheet_name="Metadata")
    append_tsv_as_sheet(xlsx_path, str(stats_tsv), sheet_name="Statistics")
    append_tsv_as_sheet(xlsx_path, str(gene_burden_tsv), sheet_name="Gene Burden")
    finalize_excel_file(xlsx_path, cfg)

    # Verify all sheets exist
    wb = load_workbook(xlsx_path)
    sheet_names = wb.sheetnames
    assert len(sheet_names) == 4, f"Expected 4 sheets, got {len(sheet_names)}"
    assert "Results" in sheet_names
    assert "Metadata" in sheet_names
    assert "Statistics" in sheet_names
    assert "Gene Burden" in sheet_names

    # Verify sheet order (Results should be first)
    assert sheet_names[0] == "Results"

    # Verify data row counts (header + data rows)
    results_ws = wb["Results"]
    metadata_ws = wb["Metadata"]
    stats_ws = wb["Statistics"]
    gene_burden_ws = wb["Gene Burden"]

    # Results: 20 data rows + 1 header
    assert results_ws.max_row == 21, f"Expected 21 rows in Results, got {results_ws.max_row}"
    # Metadata: 2 data rows + 1 header
    assert metadata_ws.max_row == 3, f"Expected 3 rows in Metadata, got {metadata_ws.max_row}"
    # Statistics: 3 data rows + 1 header
    assert stats_ws.max_row == 4, f"Expected 4 rows in Statistics, got {stats_ws.max_row}"
    # Gene Burden: 3 data rows + 1 header
    assert (
        gene_burden_ws.max_row == 4
    ), f"Expected 4 rows in Gene Burden, got {gene_burden_ws.max_row}"


@pytest.mark.unit
def test_full_excel_freeze_panes_all_sheets(tmp_path):
    """
    Verify freeze_panes is set to A2 on ALL sheets.

    After finalize_excel_file, every sheet should have the top row frozen.
    """
    # Create minimal test data
    variants_df = _create_test_variants_df(n_variants=10)
    metadata_df = pd.DataFrame({"Key": ["test"], "Value": ["value"]})

    # Write TSVs
    variants_tsv = tmp_path / "variants.tsv"
    metadata_tsv = tmp_path / "metadata.tsv"
    variants_df.to_csv(variants_tsv, sep="\t", index=False)
    metadata_df.to_csv(metadata_tsv, sep="\t", index=False)

    # Generate Excel with 2 sheets
    cfg = {"links": {}, "igv_enabled": False}
    xlsx_path = convert_to_excel(str(variants_tsv), cfg, df=variants_df)
    append_tsv_as_sheet(xlsx_path, str(metadata_tsv), sheet_name="Metadata")
    finalize_excel_file(xlsx_path, cfg)

    # Verify freeze panes on all sheets
    wb = load_workbook(xlsx_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.freeze_panes == "A2", (
            f"Expected freeze_panes='A2' on sheet '{sheet_name}', got '{ws.freeze_panes}'"
        )


@pytest.mark.unit
def test_full_excel_auto_filter_all_sheets(tmp_path):
    """
    Verify auto_filter is set on ALL sheets.

    After finalize_excel_file, every sheet should have auto-filter enabled.
    """
    # Create minimal test data
    variants_df = _create_test_variants_df(n_variants=10)
    metadata_df = pd.DataFrame({"Key": ["test"], "Value": ["value"]})

    # Write TSVs
    variants_tsv = tmp_path / "variants.tsv"
    metadata_tsv = tmp_path / "metadata.tsv"
    variants_df.to_csv(variants_tsv, sep="\t", index=False)
    metadata_df.to_csv(metadata_tsv, sep="\t", index=False)

    # Generate Excel with 2 sheets
    cfg = {"links": {}, "igv_enabled": False}
    xlsx_path = convert_to_excel(str(variants_tsv), cfg, df=variants_df)
    append_tsv_as_sheet(xlsx_path, str(metadata_tsv), sheet_name="Metadata")
    finalize_excel_file(xlsx_path, cfg)

    # Verify auto-filter on all sheets
    wb = load_workbook(xlsx_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Check that auto_filter.ref is set
        assert ws.auto_filter.ref is not None, (
            f"Expected auto_filter to be set on sheet '{sheet_name}', but auto_filter.ref is None"
        )
        # Should be in format "A1:X1" where X is the last column
        assert ws.auto_filter.ref.startswith("A1:"), (
            f"Expected auto_filter.ref to start with 'A1:' on sheet '{sheet_name}', "
            f"got '{ws.auto_filter.ref}'"
        )


@pytest.mark.unit
def test_full_excel_url_hyperlinks(tmp_path):
    """
    Verify URL columns are converted to clickable hyperlinks with proper styling.

    finalize_excel_file should:
    - Detect URL columns (>70% cells start with http)
    - Convert cell values to hyperlinks
    - Apply Hyperlink style
    - Use column name as display text
    """
    # Create test data with URL columns
    variants_df = _create_test_variants_df(n_variants=20)

    # Write TSV
    variants_tsv = tmp_path / "variants.tsv"
    variants_df.to_csv(variants_tsv, sep="\t", index=False)

    # Generate Excel
    cfg = {"links": {}, "igv_enabled": False}
    xlsx_path = convert_to_excel(str(variants_tsv), cfg, df=variants_df)
    finalize_excel_file(xlsx_path, cfg)

    # Load and verify hyperlinks
    wb = load_workbook(xlsx_path)
    ws = wb["Results"]

    # Find URL column indices
    header_row = [cell.value for cell in ws[1]]
    splicai_col_idx = header_row.index("SpliceAI_URL") + 1 if "SpliceAI_URL" in header_row else None
    gnomad_col_idx = header_row.index("gnomAD_URL") + 1 if "gnomAD_URL" in header_row else None

    assert splicai_col_idx is not None, "SpliceAI_URL column not found"
    assert gnomad_col_idx is not None, "gnomAD_URL column not found"

    # Check hyperlinks in data rows
    hyperlink_count = 0
    total_checked = 0

    for row_idx in range(2, min(12, ws.max_row + 1)):  # Check first 10 data rows
        # Check SpliceAI_URL column
        splicai_cell = ws.cell(row=row_idx, column=splicai_col_idx)
        if splicai_cell.hyperlink:
            hyperlink_count += 1
            # Verify hyperlink points to a URL
            target = str(splicai_cell.hyperlink.target)
            assert target.startswith("https://"), (
                f"Expected hyperlink to start with 'https://', got '{target}'"
            )
            # Verify style is set
            assert splicai_cell.style == "Hyperlink", (
                f"Expected style='Hyperlink', got '{splicai_cell.style}'"
            )
        total_checked += 1

        # Check gnomAD_URL column
        gnomad_cell = ws.cell(row=row_idx, column=gnomad_col_idx)
        if gnomad_cell.hyperlink:
            hyperlink_count += 1
        total_checked += 1

    # At least 70% of checked cells should have hyperlinks (matching converter.py threshold)
    hyperlink_ratio = hyperlink_count / total_checked if total_checked > 0 else 0
    assert hyperlink_ratio >= 0.7, (
        f"Expected >= 70% of URL cells to have hyperlinks, got {hyperlink_ratio * 100:.1f}%"
    )


@pytest.mark.unit
def test_gt_cache_cleanup_before_output(tmp_path):
    """
    Verify cache column cleanup pattern works correctly.

    Tests the cache cleanup pattern used in output stages:
    Drop all columns starting with "_" before writing to Excel.

    NOTE: convert_to_excel itself doesn't strip cache columns - that's
    the responsibility of the calling stage (TSVOutputStage, ExcelReportStage).
    This test verifies the cleanup pattern works correctly.
    """
    # Create test data with cache columns
    variants_df = _create_test_variants_df(n_variants=10)

    # Add cache columns (simulating what dataframe_optimizer adds)
    variants_df["_GT_PARSED"] = [
        [{"sample": f"Sample{i}", "gt": "0/1"}] for i in range(len(variants_df))
    ]
    variants_df["_INTERNAL_CACHE"] = ["cache_value"] * len(variants_df)

    # Apply the cache cleanup pattern from output stages
    cache_cols = [c for c in variants_df.columns if c.startswith("_")]
    assert len(cache_cols) == 2, f"Expected 2 cache columns, found {len(cache_cols)}"

    df_cleaned = variants_df.drop(columns=cache_cols)

    # Write TSV and generate Excel with cleaned DataFrame
    variants_tsv = tmp_path / "variants.tsv"
    df_cleaned.to_csv(variants_tsv, sep="\t", index=False)

    cfg = {"links": {}, "igv_enabled": False}
    xlsx_path = convert_to_excel(str(variants_tsv), cfg, df=df_cleaned)
    finalize_excel_file(xlsx_path, cfg)

    # Verify no cache columns in Excel
    wb = load_workbook(xlsx_path)
    ws = wb["Results"]
    header_row = [cell.value for cell in ws[1]]

    # Verify all expected columns are present
    expected_cols = [
        "CHROM", "POS", "REF", "ALT", "GT", "GENE", "IMPACT", "SpliceAI_URL", "gnomAD_URL"
    ]
    for col in expected_cols:
        assert col in header_row, f"Expected column '{col}' missing from output"

    # Verify no cache columns (starting with _)
    cache_cols_in_output = [c for c in header_row if c and c.startswith("_")]
    assert len(cache_cols_in_output) == 0, (
        f"Cache columns should not appear in Excel output. Found: {cache_cols_in_output}"
    )
