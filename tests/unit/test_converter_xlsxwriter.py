"""
Tests for xlsxwriter-based Excel generation in converter.py.

These tests verify that the two-pass Excel generation (xlsxwriter for bulk write,
openpyxl for finalization) produces functionally equivalent output to the original
openpyxl-only approach.
"""

import os
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from variantcentrifuge.converter import (
    append_tsv_as_sheet,
    convert_to_excel,
    finalize_excel_file,
)


@pytest.mark.unit
def test_convert_to_excel_uses_xlsxwriter(tmp_path: Path) -> None:
    """Test that convert_to_excel creates valid Excel file using xlsxwriter."""
    # Create a small synthetic DataFrame
    df = pd.DataFrame(
        {
            "CHROM": ["chr1", "chr2", "chr3"],
            "POS": [100, 200, 300],
            "REF": ["A", "C", "G"],
            "ALT": ["T", "G", "A"],
            "GENE": ["GENE1", "GENE2", "GENE3"],
            "IMPACT": ["HIGH", "MODERATE", "LOW"],
        }
    )

    # Write to TSV
    tsv_file = tmp_path / "test_variants.tsv"
    df.to_csv(tsv_file, sep="\t", index=False)

    # Convert to Excel (using xlsxwriter engine internally)
    cfg = {"links": {}}
    xlsx_file = convert_to_excel(str(tsv_file), cfg)

    # Verify Excel file was created
    assert os.path.exists(xlsx_file)
    assert xlsx_file.endswith(".xlsx")

    # Open with openpyxl and verify contents
    wb = load_workbook(xlsx_file)
    assert "Results" in wb.sheetnames
    ws = wb["Results"]

    # Verify column headers
    header_row = [cell.value for cell in ws[1]]
    assert header_row == ["CHROM", "POS", "REF", "ALT", "GENE", "IMPACT"]

    # Verify data rows
    assert ws.max_row == 4  # Header + 3 data rows
    assert ws.cell(row=2, column=1).value == "chr1"
    # POS is read as string due to genomic_dtypes in convert_to_excel
    assert str(ws.cell(row=2, column=2).value) == "100"
    assert ws.cell(row=3, column=5).value == "GENE2"

    wb.close()


@pytest.mark.unit
def test_convert_to_excel_with_dataframe(tmp_path: Path) -> None:
    """Test convert_to_excel with in-memory DataFrame (skips disk read)."""
    df = pd.DataFrame(
        {
            "CHROM": ["chr1"],
            "POS": [12345],
            "REF": ["G"],
            "ALT": ["T"],
        }
    )

    # Create dummy TSV path (won't be read)
    tsv_file = tmp_path / "dummy.tsv"
    tsv_file.write_text("CHROM\tPOS\n")  # Dummy content

    # Convert using provided DataFrame
    cfg = {"links": {}}
    xlsx_file = convert_to_excel(str(tsv_file), cfg, df=df)

    # Verify output
    wb = load_workbook(xlsx_file)
    ws = wb["Results"]
    assert ws.cell(row=2, column=1).value == "chr1"
    # POS is read as string due to genomic_dtypes in convert_to_excel
    assert str(ws.cell(row=2, column=2).value) == "12345"
    wb.close()


@pytest.mark.unit
def test_finalize_adds_freeze_panes(tmp_path: Path) -> None:
    """Test that finalize_excel_file adds freeze panes to all sheets."""
    df = pd.DataFrame(
        {
            "CHROM": ["chr1", "chr2"],
            "POS": [100, 200],
        }
    )

    tsv_file = tmp_path / "test.tsv"
    df.to_csv(tsv_file, sep="\t", index=False)

    cfg = {"links": {}}
    xlsx_file = convert_to_excel(str(tsv_file), cfg)
    finalize_excel_file(xlsx_file, cfg)

    wb = load_workbook(xlsx_file)
    for ws in wb.worksheets:
        assert ws.freeze_panes == "A2", f"Sheet {ws.title} should have freeze panes at A2"
    wb.close()


@pytest.mark.unit
def test_finalize_adds_auto_filter(tmp_path: Path) -> None:
    """Test that finalize_excel_file adds auto-filters to all sheets."""
    df = pd.DataFrame(
        {
            "CHROM": ["chr1"],
            "POS": [100],
            "REF": ["A"],
            "ALT": ["T"],
        }
    )

    tsv_file = tmp_path / "test.tsv"
    df.to_csv(tsv_file, sep="\t", index=False)

    cfg = {"links": {}}
    xlsx_file = convert_to_excel(str(tsv_file), cfg)
    finalize_excel_file(xlsx_file, cfg)

    wb = load_workbook(xlsx_file)
    ws = wb["Results"]

    # Verify auto-filter is set
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith("A1:")
    assert "D1" in ws.auto_filter.ref  # Should span all 4 columns
    wb.close()


@pytest.mark.unit
def test_finalize_adds_hyperlinks(tmp_path: Path) -> None:
    """Test that finalize_excel_file adds hyperlinks to URL columns."""
    df = pd.DataFrame(
        {
            "CHROM": ["chr1", "chr2"],
            "POS": [100, 200],
            "SpliceAI": [
                "https://spliceailookup.broadinstitute.org/variant1",
                "https://spliceailookup.broadinstitute.org/variant2",
            ],
            "Franklin": [
                "https://franklin.genoox.com/variant1",
                "https://franklin.genoox.com/variant2",
            ],
        }
    )

    tsv_file = tmp_path / "test.tsv"
    df.to_csv(tsv_file, sep="\t", index=False)

    # Create config with link templates (not strictly needed for this test
    # since we're directly checking URL columns, but matches real usage)
    cfg = {
        "links": {
            "SpliceAI": "https://spliceailookup.broadinstitute.org/{CHROM}_{POS}",
            "Franklin": "https://franklin.genoox.com/{CHROM}_{POS}",
        }
    }

    xlsx_file = convert_to_excel(str(tsv_file), cfg)
    finalize_excel_file(xlsx_file, cfg)

    wb = load_workbook(xlsx_file)
    ws = wb["Results"]

    # Find SpliceAI and Franklin columns (should be columns 3 and 4)
    header_row = [cell.value for cell in ws[1]]
    spliceai_col = header_row.index("SpliceAI") + 1
    franklin_col = header_row.index("Franklin") + 1

    # Check that hyperlinks were added and styled correctly
    for row_idx in range(2, ws.max_row + 1):
        # SpliceAI column
        cell = ws.cell(row=row_idx, column=spliceai_col)
        assert cell.hyperlink is not None, f"Row {row_idx} SpliceAI should have hyperlink"
        assert cell.style == "Hyperlink", f"Row {row_idx} SpliceAI should use Hyperlink style"
        assert cell.value == "SpliceAI", f"Row {row_idx} SpliceAI should display column name"

        # Franklin column
        cell = ws.cell(row=row_idx, column=franklin_col)
        assert cell.hyperlink is not None, f"Row {row_idx} Franklin should have hyperlink"
        assert cell.style == "Hyperlink", f"Row {row_idx} Franklin should use Hyperlink style"
        assert cell.value == "Franklin", f"Row {row_idx} Franklin should display column name"

    wb.close()


@pytest.mark.unit
def test_append_tsv_as_sheet(tmp_path: Path) -> None:
    """Test that append_tsv_as_sheet correctly adds metadata sheet."""
    # Create initial Results sheet
    df_results = pd.DataFrame(
        {
            "CHROM": ["chr1"],
            "POS": [100],
        }
    )

    tsv_file = tmp_path / "test.tsv"
    df_results.to_csv(tsv_file, sep="\t", index=False)

    cfg = {"links": {}}
    xlsx_file = convert_to_excel(str(tsv_file), cfg)

    # Create metadata TSV
    df_metadata = pd.DataFrame(
        {
            "Key": ["Version", "Date", "Samples"],
            "Value": ["0.13.0", "2026-02-15", "100"],
        }
    )

    metadata_tsv = tmp_path / "metadata.tsv"
    df_metadata.to_csv(metadata_tsv, sep="\t", index=False)

    # Append metadata sheet
    append_tsv_as_sheet(xlsx_file, str(metadata_tsv), sheet_name="Metadata")

    # Verify both sheets exist
    wb = load_workbook(xlsx_file)
    assert "Results" in wb.sheetnames
    assert "Metadata" in wb.sheetnames

    # Verify metadata content
    ws_metadata = wb["Metadata"]
    assert ws_metadata.cell(row=1, column=1).value == "Key"
    assert ws_metadata.cell(row=1, column=2).value == "Value"
    assert ws_metadata.cell(row=2, column=1).value == "Version"
    assert ws_metadata.cell(row=2, column=2).value == "0.13.0"

    wb.close()


@pytest.mark.unit
def test_empty_dataframe_handling(tmp_path: Path) -> None:
    """Test that empty DataFrame produces valid Excel with headers only."""
    # Create empty DataFrame with columns only
    df = pd.DataFrame(columns=["CHROM", "POS", "REF", "ALT", "GENE"])

    tsv_file = tmp_path / "empty.tsv"
    df.to_csv(tsv_file, sep="\t", index=False)

    cfg = {"links": {}}
    xlsx_file = convert_to_excel(str(tsv_file), cfg)

    # Verify Excel file was created with headers
    wb = load_workbook(xlsx_file)
    ws = wb["Results"]

    # Should have 1 row (header only)
    assert ws.max_row == 1
    header_row = [cell.value for cell in ws[1]]
    assert header_row == ["CHROM", "POS", "REF", "ALT", "GENE"]

    wb.close()


@pytest.mark.unit
def test_finalize_with_empty_dataframe(tmp_path: Path) -> None:
    """Test that finalize_excel_file works correctly on empty Excel files."""
    df = pd.DataFrame(columns=["CHROM", "POS"])

    tsv_file = tmp_path / "empty.tsv"
    df.to_csv(tsv_file, sep="\t", index=False)

    cfg = {"links": {}}
    xlsx_file = convert_to_excel(str(tsv_file), cfg)
    finalize_excel_file(xlsx_file, cfg)

    # Verify freeze panes and auto-filter are still added
    wb = load_workbook(xlsx_file)
    ws = wb["Results"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    wb.close()


@pytest.mark.unit
def test_igv_links_column_removed(tmp_path: Path) -> None:
    """Test that raw igv_links column is removed from Excel output."""
    df = pd.DataFrame(
        {
            "CHROM": ["chr1"],
            "POS": [100],
            "igv_links": ["igv/sample1.html"],  # Raw IGV links column
        }
    )

    tsv_file = tmp_path / "test.tsv"
    df.to_csv(tsv_file, sep="\t", index=False)

    cfg = {"links": {}}
    xlsx_file = convert_to_excel(str(tsv_file), cfg)

    wb = load_workbook(xlsx_file)
    ws = wb["Results"]

    # Verify igv_links column was removed
    header_row = [cell.value for cell in ws[1]]
    assert "igv_links" not in header_row
    assert "CHROM" in header_row
    assert "POS" in header_row

    wb.close()


@pytest.mark.unit
def test_xlsxwriter_openpyxl_compatibility(tmp_path: Path) -> None:
    """Test that openpyxl can read and modify xlsxwriter-generated files."""
    # Create Excel with xlsxwriter
    df = pd.DataFrame(
        {
            "CHROM": ["chr1", "chr2"],
            "POS": [100, 200],
            "REF": ["A", "C"],
        }
    )

    tsv_file = tmp_path / "test.tsv"
    df.to_csv(tsv_file, sep="\t", index=False)

    cfg = {"links": {}}
    xlsx_file = convert_to_excel(str(tsv_file), cfg)

    # Load with openpyxl and modify
    wb = load_workbook(xlsx_file)
    ws = wb["Results"]

    # Modify a cell
    ws.cell(row=2, column=1).value = "chrX"

    # Add a new sheet
    ws2 = wb.create_sheet("TestSheet")
    ws2["A1"] = "Test"

    # Save modifications
    wb.save(xlsx_file)
    wb.close()

    # Reload and verify modifications persisted
    wb2 = load_workbook(xlsx_file)
    assert wb2["Results"].cell(row=2, column=1).value == "chrX"
    assert "TestSheet" in wb2.sheetnames
    assert wb2["TestSheet"]["A1"].value == "Test"
    wb2.close()
