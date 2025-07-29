#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Script to create a cohort report from multiple sample TSV files.

create_cohort_report.py

This script aggregates results from multiple single-sample VariantCentrifuge reports
into one comprehensive, interactive cohort-level report.

Author: VariantCentrifuge Team
"""

import argparse
import glob
import json
import logging
import os
import re
import sys

import pandas as pd
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger("create_cohort_report")


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Create a cohort report from multiple VariantCentrifuge sample reports."
    )

    parser.add_argument(
        "--input-pattern",
        required=True,
        help='Glob pattern to find input TSV files (e.g., "/path/to/samples/*/variants.tsv")',
    )
    parser.add_argument(
        "--output-dir", required=True, help="Directory where the cohort report will be saved"
    )
    parser.add_argument(
        "--report-name",
        default="Variant Cohort Analysis",
        help='Name for the cohort report (default: "Variant Cohort Analysis")',
    )
    parser.add_argument(
        "--sample-regex",
        default=r".*?([^/\\]+)(?:/|\\)variants\.tsv$",
        help="Regex pattern to extract sample ID from input file paths",
    )

    return parser.parse_args()


def find_input_files(input_pattern):
    """Find all input TSV files using the provided glob pattern."""
    files = glob.glob(input_pattern, recursive=True)

    if not files:
        logger.error(f"No input files found matching pattern: {input_pattern}")
        sys.exit(1)

    logger.info(f"Found {len(files)} input files")
    return files


def extract_sample_id(file_path, regex_pattern):
    """Extract sample ID from file path using the provided regex pattern."""
    logger.debug(
        f"Attempting to extract sample ID from: '{file_path}' using regex: '{regex_pattern}'"
    )
    match = re.search(regex_pattern, file_path)
    if match:
        # Ensure group 1 exists and is not None
        if len(match.groups()) > 0 and match.group(1):
            sample_id = match.group(1)
            logger.info(
                f"Regex match successful. Extracted SampleID: '{sample_id}' from '{file_path}'"
            )
            return sample_id
        else:
            # This case handles if regex matches but capture group 1 is missing or empty
            logger.warning(
                f"Regex matched for '{file_path}', but capture group 1 was missing or empty. "
                f"Match groups: {match.groups()}. Regex: '{regex_pattern}'"
            )
    else:
        logger.warning(
            f"Regex pattern '{regex_pattern}' did not match for file path: '{file_path}'."
        )

    # Fallback: use the name of the parent directory of the file_path
    # This is often more robust if 'variants.tsv' (or similar) is consistently named
    # and located within a sample-specific directory.
    try:
        parent_dir_name = os.path.basename(os.path.dirname(file_path))
        if parent_dir_name:  # Ensure parent_dir_name is not empty
            logger.info(
                f"Using fallback: parent directory name '{parent_dir_name}' as SampleID "
                f"for '{file_path}'"
            )
            return parent_dir_name
        else:  # Handle cases like file_path being in root directory, e.g. "/variants.tsv"
            logger.warning(
                f"Fallback to parent directory name failed for '{file_path}' "
                f"(parent_dir_name is empty). Using filename as last resort."
            )
    except Exception as e:
        logger.error(f"Error during fallback parent directory extraction for '{file_path}': {e}")

    # Last resort fallback: filename without extension
    final_fallback_id = os.path.splitext(os.path.basename(file_path))[0]
    logger.info(
        f"Using last resort fallback: filename without extension '{final_fallback_id}' "
        f"as SampleID for '{file_path}'"
    )
    return final_fallback_id


def aggregate_data(input_files, sample_regex):
    """Aggregate data from multiple TSV files into a single DataFrame."""
    logger.info(
        f"Starting data aggregation from {len(input_files)} input file(s) using pattern "
        f"for sample_regex: '{sample_regex}'"
    )

    dfs = []
    collected_sample_ids = []  # To track all extracted IDs

    for file_path in input_files:
        logger.info(f"Processing file: {file_path}")  # More prominent log for each file
        try:
            # Use low_memory=False to suppress dtype warnings for variant files
            df = pd.read_csv(file_path, sep="\t", low_memory=False)
            logger.debug(f"Successfully read {file_path}. Columns: {df.columns.tolist()}")

            sample_id = extract_sample_id(
                file_path, sample_regex
            )  # Calls the updated extract_sample_id
            df["SampleID"] = sample_id
            collected_sample_ids.append(sample_id)
            logger.info(f"Assigned SampleID '{sample_id}' to data from {file_path}")

            # Ensure the Gene column exists (case-insensitive check and rename)
            if "Gene" not in df.columns:
                found_gene_col_original_case = None
                for col_idx, col_name in enumerate(df.columns):
                    if col_name.lower() == "gene":
                        found_gene_col_original_case = col_name
                        break
                if found_gene_col_original_case:
                    df.rename(columns={found_gene_col_original_case: "Gene"}, inplace=True)
                    logger.debug(
                        f"Renamed column '{found_gene_col_original_case}' to 'Gene' "
                        f"for sample '{sample_id}' from '{file_path}'"
                    )
                else:
                    logger.warning(
                        f"'Gene' column (or case-insensitive 'gene') not found in {file_path} "
                        f"for sample '{sample_id}'. Adding 'Unknown' as placeholder."
                    )
                    df["Gene"] = "Unknown"  # Add placeholder if no gene column

            dfs.append(df)

        except Exception as e:
            logger.error(
                f"Failed to process {file_path}. Error: {str(e)}", exc_info=True
            )  # exc_info=True for traceback

    if not dfs:
        logger.critical(
            "CRITICAL: No dataframes were created after processing all input files. "
            "This means no valid data could be read or processed. Exiting."
        )
        sys.exit(1)

    master_df = pd.concat(dfs, ignore_index=True)

    unique_extracted_samples = sorted(
        list(set(collected_sample_ids))
    )  # Get sorted list of unique IDs
    logger.info(f"Aggregation complete. Total variants aggregated: {len(master_df)}.")
    logger.info(
        f"All SampleIDs extracted during aggregation (includes duplicates if any): "
        f"{collected_sample_ids}"
    )
    logger.info(
        f"Number of unique SampleIDs successfully extracted: {len(unique_extracted_samples)}. "
        f"Unique IDs found: {unique_extracted_samples}"
    )

    if "SampleID" in master_df.columns:
        df_unique_samples_count = master_df["SampleID"].nunique()
        df_unique_samples_list = sorted(list(master_df["SampleID"].unique()))
        logger.info(
            f"Verification: Unique SampleIDs present in the final aggregated DataFrame: "
            f"{df_unique_samples_count}. IDs: {df_unique_samples_list}"
        )
        if (
            len(unique_extracted_samples) != df_unique_samples_count
            or unique_extracted_samples != df_unique_samples_list
        ):
            logger.warning(
                "Potential Mismatch Alert! The set of unique SampleIDs extracted during "
                "file processing "
                f"({len(unique_extracted_samples)}: {unique_extracted_samples}) "
                "differs from the unique SampleIDs found in the final DataFrame "
                f"({df_unique_samples_count}: {df_unique_samples_list}). "
                "This could indicate an issue in data concatenation or SampleID assignment."
            )
    else:
        logger.error(
            "CRITICAL: 'SampleID' column is MISSING from the final aggregated DataFrame. "
            "Sample identification will fail."
        )

    return master_df


def clean_data(df):
    """Clean and prepare the aggregated data."""
    logger.info("Cleaning and preparing data...")

    # Ensure SampleID is the first column
    if "SampleID" in df.columns:
        # Get all columns except SampleID
        other_columns = [col for col in df.columns if col != "SampleID"]
        # Reorder columns with SampleID first
        column_order = ["SampleID"] + other_columns
        df = df[column_order]
        logger.info("Reordered columns with SampleID as first column")

    # Ensure all relevant numeric columns are properly typed using pattern-based detection
    logger.info("Converting numeric columns to proper data types...")

    for col in df.columns:
        # Pattern-based detection for columns that should be numeric
        should_convert = False

        # AF columns (allele frequency): AF_1, AF_EXAC, AF_GNOMAD, etc.
        if col.startswith("AF_") or col == "AF":
            should_convert = True
        # AD columns (allelic depth): AD_1, AD_2, etc.
        elif col.startswith("AD_") or col == "AD":
            should_convert = True
        # Common numeric columns
        elif col in ["QUAL", "DP", "IMPACT_SEVERITY", "CADD_PHRED"]:
            should_convert = True
        # Other common numeric patterns
        elif any(col.endswith(suffix) for suffix in ["_score", "_SCORE", "_phred", "_PHRED"]):
            should_convert = True

        if should_convert:
            original_dtype = df[col].dtype
            df[col] = pd.to_numeric(df[col], errors="coerce")
            logger.debug(f"Converted column '{col}' from {original_dtype} to numeric")

    # Clean up any NaN values in string columns
    string_cols = ["Gene", "GeneID", "IMPACT", "Effect", "AAChange", "SampleID"]
    for col in string_cols:
        if col in df.columns:
            df[col] = df[col].fillna("Unknown")

    # Sort by gene and impact severity
    if "Gene" in df.columns:
        df = df.sort_values(by=["Gene", "SampleID"])

    return df


def compute_statistics(df):
    """Compute summary statistics for the cohort."""
    logger.info("Computing summary statistics...")

    # Global summary
    total_variants = len(df)
    unique_samples = df["SampleID"].nunique()
    unique_genes = df["Gene"].nunique() if "Gene" in df.columns else 0

    # Per-gene statistics
    gene_summary = None
    if "Gene" in df.columns:
        # Count variants per gene
        gene_variants = df.groupby("Gene").size().reset_index(name="VariantCount")

        # Count unique samples per gene
        gene_samples = df.groupby("Gene")["SampleID"].nunique().reset_index(name="SampleCount")

        # Merge the two statistics
        gene_summary = pd.merge(gene_variants, gene_samples, on="Gene")

        # Sort by number of samples and then by number of variants
        gene_summary = gene_summary.sort_values(by=["SampleCount", "VariantCount"], ascending=False)

    summary = {
        "total_variants": int(total_variants),
        "unique_samples": int(unique_samples),
        "unique_genes": int(unique_genes),
        "gene_summary": gene_summary.to_dict(orient="records") if gene_summary is not None else [],
    }

    return summary


def save_data(df, summary, output_dir):
    """Save the prepared data as JSON files."""
    logger.info("Saving data to JSON files...")

    # Create data directory if it doesn't exist
    data_dir = os.path.join(output_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    # Save variants data
    variants_file = os.path.join(data_dir, "variants.json")
    with open(variants_file, "w") as f:
        # Convert to JSON records with date_format='iso' for proper datetime handling
        variants_json = df.to_json(orient="records", date_format="iso")
        f.write(variants_json)

    # Save summary data
    summary_file = os.path.join(data_dir, "summary.json")
    with open(summary_file, "w") as f:
        json.dump(summary, f, indent=2)

    logger.info(f"Data saved to {data_dir}")


def detect_link_columns(df):
    """Detect columns that contain URLs or should be treated as links."""
    link_columns = {}

    if df.empty:
        return link_columns

    for col in df.columns:
        if col.lower() in ["url", "link", "href"]:
            link_columns[col] = {"type": "standard", "display_text": "Link"}
            continue

        # Check if column contains URLs
        sample_values = df[col].dropna().astype(str).head(10)
        url_count = sum(
            1 for val in sample_values if val.startswith(("http://", "https://", "ftp://"))
        )

        # If more than 50% of non-null values are URLs, treat as link column
        if len(sample_values) > 0 and url_count / len(sample_values) > 0.5:
            link_columns[col] = {"type": "standard", "display_text": col.replace("_", " ")}

    logger.info(f"Detected link columns: {list(link_columns.keys())}")
    return link_columns


def prepare_column_metadata(df):
    """Prepare column metadata for template rendering."""
    link_columns = detect_link_columns(df)

    # Use the same default hidden columns as individual reports from config.json
    default_hidden_columns = [
        "QUAL",
        "AC",
        "FEATUREID",
        "AA_POS",
        "AA_LEN",
        "NMD_PERC",
        "dbNSFP_ALFA_Total_AC",
        "dbNSFP_REVEL_score",
        "splice_dbscSNV_rf_score",
        "splice_dbscSNV_ada_score",
        "splice_spidex_dpsi_zscore",
        "dbNSFP_clinvar_clnsig",
        "dbNSFP_gnomAD_exomes_AC",
        "dbNSFP_gnomAD_genomes_AC",
        "hgmd_CLASS",
        "proband_count",
        "control_count",
        "proband_variant_count",
        "control_variant_count",
        "proband_allele_count",
        "control_allele_count",
        "proband_homozygous_count",
        "control_homozygous_count",
    ]

    # Use the same hover-expand columns as individual reports from config.json
    truncate_columns = [
        "VAR_ID",
        "HGVS_P",
        "HGVS_C",
        "ID",
        "REF",
        "ALT",
        "EFFECT",
        "IMPACT",
        "FEATUREID",
        "dbNSFP_clinvar_clnsig",
        "ClinVar_CLNSIG",
        "GT",
    ]

    column_metadata = []
    for col in df.columns:
        is_link = col in link_columns
        is_hidden = col in default_hidden_columns
        needs_truncation = col in truncate_columns
        is_igv_column = col == "igv_links"

        # Special handling for IGV links column
        if is_igv_column:
            metadata = {
                "name": col,
                "display_name": "IGV Report Links",
                "is_link": False,  # Special IGV rendering, not standard link
                "is_igv_link_column": True,
                "link_info": {},
                "is_hidden_default": False,  # Show IGV links by default
                "needs_truncation": False,
                "max_width": None,
            }
        else:
            metadata = {
                "name": col,
                "display_name": col.replace("_", " "),
                "is_link": is_link,
                "is_igv_link_column": False,
                "link_info": link_columns.get(col, {}),
                "is_hidden_default": is_hidden,
                "needs_truncation": needs_truncation,
                "max_width": 250 if col == "GT" else (120 if needs_truncation else None),
            }
        column_metadata.append(metadata)

    return column_metadata


def discover_igv_maps(input_files, output_dir):
    """Discover IGV report maps from individual sample directories."""
    logger.info("Discovering IGV report maps from sample directories...")

    igv_maps = []
    igv_lookup = {}

    for file_path in input_files:
        try:
            # Try to find IGV map in the same directory structure as the sample
            sample_dir = os.path.dirname(file_path)

            # Look for IGV map in common locations
            possible_igv_paths = [
                os.path.join(sample_dir, "report", "igv", "igv_reports_map.json"),
                os.path.join(sample_dir, "output", "report", "igv", "igv_reports_map.json"),
                os.path.join(os.path.dirname(sample_dir), "report", "igv", "igv_reports_map.json"),
            ]

            igv_map_path = None
            igv_base_dir = None
            for path in possible_igv_paths:
                if os.path.exists(path):
                    igv_map_path = path
                    # Store the base directory where IGV reports are located
                    igv_base_dir = os.path.dirname(
                        os.path.dirname(path)
                    )  # Remove /igv/igv_reports_map.json
                    break

            if igv_map_path:
                logger.info(f"Found IGV map: {igv_map_path}")

                with open(igv_map_path, "r", encoding="utf-8") as f:
                    igv_map_data = json.load(f)

                # Handle both old and new IGV map format
                if "variants" in igv_map_data:
                    igv_map = igv_map_data["variants"]
                else:
                    igv_map = igv_map_data

                igv_maps.extend(igv_map)

                # Populate lookup dictionary for quick access
                for entry in igv_map:
                    chrom = entry.get("chrom", "")
                    pos = str(entry.get("pos", ""))
                    ref = entry.get("ref", "")
                    alt = entry.get("alt", "")

                    # Handle new format with sample_reports nested dictionary
                    if "sample_reports" in entry:
                        for sample_id, report_path in entry["sample_reports"].items():
                            key = (chrom, pos, ref, alt, sample_id)
                            # Create absolute path to the IGV report
                            absolute_igv_path = os.path.join(igv_base_dir, report_path)
                            # Calculate relative path from cohort output directory to IGV report
                            relative_path = os.path.relpath(absolute_igv_path, output_dir)
                            igv_lookup[key] = relative_path
                    # Handle old format with flat sample_id and report_path
                    elif "sample_id" in entry and "report_path" in entry:
                        key = (chrom, pos, ref, alt, entry["sample_id"])
                        # Create absolute path to the IGV report
                        absolute_igv_path = os.path.join(igv_base_dir, entry["report_path"])
                        # Calculate relative path from cohort output directory to IGV report
                        relative_path = os.path.relpath(absolute_igv_path, output_dir)
                        igv_lookup[key] = relative_path

        except Exception as e:
            logger.debug(f"No IGV map found or error loading from {file_path}: {e}")
            continue

    if igv_maps:
        logger.info(f"Loaded {len(igv_maps)} IGV map entries from {len(igv_lookup)} sample reports")
        logger.info("IGV paths resolved relative to cohort output directory")
    else:
        logger.info("No IGV maps found - cohort report will not include IGV links")

    return igv_lookup


def enrich_variants_with_igv(df, igv_lookup):
    """Enrich variants DataFrame with IGV links using the same logic as individual reports."""
    if not igv_lookup:
        logger.info("No IGV lookup available - skipping IGV enrichment")
        df["igv_links"] = [[] for _ in range(len(df))]
        return df

    logger.info("Enriching variants with IGV report links...")

    # Pattern to extract (SampleID, GenotypeString) tuples from GT column
    pattern = re.compile(r"([^()]+)\(([^)]+)\)")

    igv_links_list = []

    for _, row in df.iterrows():
        igv_links = []

        # Extract variant identifiers
        chrom = str(row.get("CHROM", ""))
        pos = str(row.get("POS", ""))
        ref = str(row.get("REF", ""))
        alt = str(row.get("ALT", ""))

        # Extract sample IDs with non-reference genotypes from GT column
        gt_value = row.get("GT", "")
        if gt_value:
            sample_entries = str(gt_value).split(";")
            for entry in sample_entries:
                entry = entry.strip()
                if not entry:
                    continue

                m = pattern.match(entry)
                if m:
                    sample_id = m.group(1).strip()
                    genotype = m.group(2).strip()

                    # Skip reference genotypes
                    if genotype in ["0/0", "./."]:
                        continue

                    # Look up the IGV report path
                    lookup_key = (chrom, pos, ref, alt, sample_id)
                    if lookup_key in igv_lookup:
                        igv_links.append(
                            {"sample_id": sample_id, "report_path": igv_lookup[lookup_key]}
                        )

        igv_links_list.append(igv_links)

    df["igv_links"] = igv_links_list

    variants_with_igv = sum(1 for links in igv_links_list if links)
    logger.info(f"Added IGV links to {variants_with_igv} variants")

    return df


def create_excel_report(df, output_dir, igv_lookup=None):
    """Create an Excel report with clickable links."""
    logger.info("Creating Excel report...")

    # Create Excel file path
    excel_file = os.path.join(output_dir, "cohort_variants.xlsx")

    # Write DataFrame to Excel
    df.to_excel(excel_file, index=False, sheet_name="Cohort_Variants")

    # Add formatting and clickable links
    finalize_cohort_excel(excel_file, df, igv_lookup)

    logger.info(f"Excel report created: {excel_file}")
    return excel_file


def finalize_cohort_excel(xlsx_file, df, igv_lookup=None):
    """Apply basic Excel formatting and add IGV Report Links column if available."""
    wb = load_workbook(xlsx_file)
    ws = wb.active

    # Freeze the top row
    ws.freeze_panes = "A2"

    # Enable auto-filter
    max_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{max_col_letter}1"

    logger.info("Finalizing Excel file with basic formatting...")

    # Get header information
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_indices = {}
    url_columns = []

    for idx, col_name in enumerate(header_row, 1):  # 1-indexed for openpyxl
        if col_name in ["CHROM", "POS", "REF", "ALT", "GT"]:
            header_indices[col_name] = idx

        # Check if this looks like a URL column
        if col_name and ws.max_row > 1:
            sample_values = []
            for row_idx in range(2, min(6, ws.max_row + 1)):
                cell_value = ws.cell(row=row_idx, column=idx).value
                if cell_value and str(cell_value).strip():
                    sample_values.append(str(cell_value).strip())

            if sample_values:
                url_count = sum(
                    1 for val in sample_values if val.startswith(("http://", "https://"))
                )
                if url_count >= len(sample_values) * 0.7:
                    url_columns.append(col_name)

    if url_columns:
        logger.info(f"Found {len(url_columns)} URL columns: {url_columns}")
        logger.info("URLs left as plain text - Excel will auto-detect as clickable links")
    else:
        logger.info("No URL columns detected")

    # Add IGV Report Links column if IGV data is available
    igv_col = None
    if igv_lookup and all(col in header_indices for col in ["CHROM", "POS", "REF", "ALT", "GT"]):
        logger.info("Adding IGV Report Links column to Excel...")

        # Add the IGV Report Links column header
        igv_col = ws.max_column + 1
        igv_col_letter = get_column_letter(igv_col)
        igv_header_cell = ws[f"{igv_col_letter}1"]
        igv_header_cell.value = "IGV Report Links"

        # Pattern to extract sample IDs and genotypes from GT column
        pattern = re.compile(r"([^()]+)\(([^)]+)\)")

        # Process data rows to add IGV links
        for row_idx in range(2, ws.max_row + 1):  # Skip header
            # Extract variant identifiers from the row
            chrom = str(ws.cell(row=row_idx, column=header_indices["CHROM"]).value or "")
            pos = str(ws.cell(row=row_idx, column=header_indices["POS"]).value or "")
            ref = str(ws.cell(row=row_idx, column=header_indices["REF"]).value or "")
            alt = str(ws.cell(row=row_idx, column=header_indices["ALT"]).value or "")

            # Get the GT value for this row
            gt_cell = ws.cell(row=row_idx, column=header_indices["GT"])
            gt_value = gt_cell.value or ""
            gt_value = str(gt_value) if gt_value is not None else ""

            if not gt_value:
                # No GT value, set N/A
                igv_cell = ws.cell(row=row_idx, column=igv_col)
                igv_cell.value = "N/A"
                continue

            # Find all related IGV reports for this variant
            igv_reports = []
            sample_entries = gt_value.split(";")
            for entry in sample_entries:
                entry = entry.strip()
                if not entry:
                    continue

                m = pattern.match(entry)
                if m:
                    sample_id = m.group(1).strip()
                    genotype = m.group(2).strip()

                    # Skip reference genotypes
                    if genotype in ["0/0", "./."]:
                        continue

                    # Look up the IGV report path
                    lookup_key = (chrom, pos, ref, alt, sample_id)
                    if lookup_key in igv_lookup:
                        igv_reports.append((sample_id, igv_lookup[lookup_key]))

            # Update the IGV links cell
            igv_cell = ws.cell(row=row_idx, column=igv_col)
            if len(igv_reports) == 1:
                # Single report - create a hyperlink like the main pipeline
                sample_id, report_path = igv_reports[0]
                # Make sure the cell value is sortable (sample ID only)
                igv_cell.value = sample_id
                # Create hyperlink to the IGV report
                igv_cell.hyperlink = report_path
                igv_cell.style = "Hyperlink"
            elif len(igv_reports) > 1:
                # Multiple reports - use first sample ID as sortable value
                first_sample = igv_reports[0][0]
                num_others = len(igv_reports) - 1
                igv_cell.value = f"{first_sample} (+{num_others} others)"

                # Add comment with full details for all reports
                from openpyxl.comments import Comment

                link_details = [f"{sid}: {rpath}" for sid, rpath in igv_reports]
                igv_cell.comment = Comment("; ".join(link_details), "IGV Reports")
            else:
                igv_cell.value = "N/A"

        # Update auto-filter to include the new IGV column
        new_max_col_letter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{new_max_col_letter}1"

        logger.info(
            f"Added IGV Report Links column with "
            f"{sum(1 for lookup_key in igv_lookup)} potential links"
        )

    # Save the workbook
    wb.save(xlsx_file)
    logger.info("Excel file finalized successfully")


def create_report(output_dir, report_name, df, summary):
    """Create the HTML report using the template and data."""
    logger.info("Creating HTML report...")

    # Get the directory of the current script
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Set up Jinja2 environment
    env = Environment(loader=FileSystemLoader(os.path.join(script_dir, "templates")))
    template = env.get_template("cohort_report_template.html")

    # Prepare column metadata for template
    column_metadata = prepare_column_metadata(df)

    # Convert DataFrame and summary to JSON strings for embedding in HTML
    variants_json = df.to_json(orient="records", date_format="iso")
    summary_json = json.dumps(summary)
    column_metadata_json = json.dumps(column_metadata)

    # Render the template with embedded JSON data
    html_content = template.render(
        report_name=report_name,
        generation_date=pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        variants_json=variants_json,
        summary_json=summary_json,
        column_metadata_json=column_metadata_json,
    )

    # Write the HTML to a file
    report_file = os.path.join(output_dir, "cohort_report.html")
    with open(report_file, "w", encoding="utf-8") as f:
        f.write(html_content)

    logger.info(f"HTML report created: {report_file}")
    return report_file


def main():
    """Execute the cohort report creation."""
    # Parse command line arguments
    args = parse_arguments()

    # Create output directory if it doesn't exist
    os.makedirs(args.output_dir, exist_ok=True)

    # Find input files
    input_files = find_input_files(args.input_pattern)

    # Discover IGV maps from sample directories
    igv_lookup = discover_igv_maps(input_files, args.output_dir)

    # Aggregate data from input files
    df = aggregate_data(input_files, args.sample_regex)

    # Clean and prepare data
    df = clean_data(df)

    # Enrich variants with IGV links if available
    df = enrich_variants_with_igv(df, igv_lookup)

    # Compute statistics
    summary = compute_statistics(df)

    # Create HTML report with embedded data (no need to save external JSON files)
    report_file = create_report(args.output_dir, args.report_name, df, summary)

    # Remove igv_links column from DataFrame for Excel export (IGV links added separately in Excel)
    df_for_excel = df.copy()
    if "igv_links" in df_for_excel.columns:
        df_for_excel = df_for_excel.drop(columns=["igv_links"])

    # Create Excel report with all variants and clickable links
    excel_file = create_excel_report(df_for_excel, args.output_dir, igv_lookup)

    logger.info("Cohort report creation complete!")
    logger.info(f"HTML report available at: {report_file}")
    logger.info(f"Excel report available at: {excel_file}")
    if igv_lookup:
        logger.info(f"IGV links integrated for {len(igv_lookup)} sample-variant combinations")


if __name__ == "__main__":
    main()
